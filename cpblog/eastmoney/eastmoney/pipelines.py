# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html

import csv
import codecs
from openpyxl import Workbook
import json

class EastmoneyPipeline:
    def process_item(self, item, spider):
        return item




class InstitutionPipeline(object): 
    def __init__(self): 
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append([
            '机构名称', 
            '机构ID', 
            '机构url',
        ])  # 设置表头



    def process_item(self, item, spider):  # 工序具体内容
        line = [
            item["institution_bxcg_name"], 
            item['institution_bxcg_code'], 
            item['institution_bxcg_url'], 
        ]  # 把数据中每一项整理出来
        self.ws.append(line)  # 将数据以行的形式添加到xlsx中
        self.wb.save(r'C:\Users\Administrator\Documents\webspider\data\BXZJ\institution_info.xlsx')  

        return item


class InstitutionInfoPipeline(object):
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(
            [
                "日期",
                "机构名称",
                "机构ID",
                "持股数量(只)",
                "持股市值(元)",
                "市值变化(1日)",
                "市值变化(5日)",
                "市值变化(10日)",
             
            ]
        )
        
    
    def process_item(self, item, spider):
        line=[
            item['institution_bxcg_date_jsondata'],
            item['institution_bxcg_name_jsondata'],
            item['institution_bxcg_code_jsondata'],
            item['institution_bxcg_num_jsondata'],
            item['institution_bxcg_value_jsondata'],
            item['institution_bxcg_valuechg1_jsondata'],
            item['institution_bxcg_valuechg5_jsondata'],
            item['institution_bxcg_valuechg10_jsondata']
        ]

        self.ws.append(line)
        self.wb.save(r'C:\Users\Administrator\Documents\webspider\data\BXZJ\history_info\institution_jsondata.xlsx')
        return item



class InstitutionDetailInfoPipeline(object):
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(
            [
                "日期",
                "公司名称",
                "公司A股代码",
                "公司H股代码",
                "机构ID",
                "机构名称",
                "持股数量(股)",
                "持股比例",
                "持股市值(元)",
                "市值变化(1日)",
                "市值变化(5日)",
                "市值变化(10日)",
             
            ]
        )

 
    def process_item(self, item, spider):
        date = item['institution_bxcg_date_detail']
        line=[

            date,
            item['company_stock_hkcode_detail'],
            item['company_stock_code_detail'],
            item['company_stock_name_detail'],
            item['institution_bxcg_code_detail'],
            item['institution_bxcg_name_detail'],
            item['institution_bxcg_stock_num_detail'],
            item['institution_bxcg_stock_ratio_detail'],
            item['institution_bxcg_stock_value_detail'],
            item['institution_bxcg_valuechg1_detail'] ,
            item['institution_bxcg_valuechg5_detail'] ,
            item['institution_bxcg_valuechg10_detail'] ,
        ]

        self.ws.append(line)
        self.wb.save(f'C:\\Users\\Administrator\\Documents\\webspider\\data\\Daily_data\\{date}\\bxzj_dailyinstitution_detail_jsondata.xlsx')
        return item





class InstitutionDailyJsondataPipeline(object):
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(
            [
                "日期",
                "公司H股代码",
                "公司A股代码",
                "公司名称",
                "机构ID",
                "机构名称",
                "持股数量(股)",
                "持股比例",
                "持股市值(元)",
                "市值变化(1日)",
                "市值变化(5日)",
                "市值变化(10日)",
             
            ]
        )

 
    def process_item(self, item, spider):
        

        
        
        
        line=[

            item['institution_bxcg_date_daily'],
            item['company_stock_hkcode_daily'],
            item['company_stock_code_daily'],
            item['company_stock_name_daily'],
            item['institution_bxcg_code_daily'],
            item['institution_bxcg_name_daily'],
            item['institution_bxcg_stock_num_daily'],
            item['institution_bxcg_stock_ratio_daily'],
            item['institution_bxcg_stock_value_daily'],
            item['institution_bxcg_valuechg1_daily'] ,
            item['institution_bxcg_valuechg5_daily'] ,
            item['institution_bxcg_valuechg10_daily'] ,
        ]

        self.ws.append(line)
        self.wb.save(r'C:\Users\Administrator\Documents\webspider\data\BXZJ\Daily_info\institution_daily_info.xlsx')
        return item


class ProxiesPipeline(object):

    def process_item(self,item,spider):
        proxis_list = item["proxies"]

        with open(r'C:\Users\Administrator\Documents\webspider\proxies.txt','w',encoding="utf-8")as f:
            f.write(proxis_list)

class   BkzjPipeline(object):
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.append(["行业名称","行业code","股票代码","股票名称"])
    def process_item(self,item,spider):
        line = [

            item["bkname"],
            item["bkcode"] ,
            item['stockcode'] ,
            item['stockname'] ,
        ] 

        self.ws.append(line)
        self.wb.save(r'C:\Users\Administrator\Desktop')
        return(item)